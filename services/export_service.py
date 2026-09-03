import io
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_participants_excel(event, registrations):
    """
    Generates a professionally styled Excel workbook (.xlsx) containing
    all participants and their custom field answers.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants Roster"

    # Define styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="0F172A")
    meta_font = Font(name="Arial", size=10, italic=True, color="475569")
    regular_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title rows
    ws.append([f"FastFest Participant Report: {event.title}"])
    ws.cell(row=1, column=1).font = title_font
    
    ws.append([f"Event Date: {event.start_time.strftime('%b %d, %Y')} | Venue: {event.venue} | Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.cell(row=2, column=1).font = meta_font
    ws.append([]) # blank line

    # Prepare custom field headers
    custom_fields = event.custom_fields
    custom_field_ids = [cf.id for cf in custom_fields]

    headers = [
        "Sl. No",
        "Reg ID",
        "Student Name",
        "Roll Number",
        "Email",
        "Department",
        "Year",
        "Section",
        "Phone",
        "Registration Status",
        "Payment Status",
        "Payment Amount (₹)",
        "Attendance Status",
        "Attended Time"
    ]
    for cf in custom_fields:
        headers.append(cf.field_label)

    ws.append(headers)
    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add participant rows
    for idx, reg in enumerate(registrations, start=1):
        student = reg.student
        profile = student.student_profile
        pay = reg.payment
        att = reg.attendance

        row_data = [
            idx,
            reg.registration_code,
            student.name,
            profile.roll_number if profile else "N/A",
            student.email,
            profile.department if profile else "N/A",
            profile.year if profile else "N/A",
            profile.section if profile else "N/A",
            student.phone or "N/A",
            reg.status,
            pay.status if pay else ("FREE" if event.is_free else "PENDING"),
            event.registration_fee,
            "PRESENT" if att else "ABSENT",
            att.scanned_at.strftime('%Y-%m-%d %H:%M:%S') if att else "N/A"
        ]

        # Map custom answers
        response_map = {r.field_id: r.field_value for r in reg.custom_responses}
        for cf_id in custom_field_ids:
            row_data.append(response_map.get(cf_id, ""))

        ws.append(row_data)
        current_row_idx = header_row_idx + idx
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=current_row_idx, column=col_idx)
            c.font = regular_font
            c.border = thin_border
            if col_idx in (1, 7, 8, 10, 11, 13):
                c.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row < 4:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_participants_csv(event, registrations):
    """
    Generates CSV string format for participants.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    custom_fields = event.custom_fields
    custom_field_ids = [cf.id for cf in custom_fields]

    headers = [
        "Sl. No", "Reg ID", "Student Name", "Roll Number", "Email",
        "Department", "Year", "Section", "Phone",
        "Registration Status", "Payment Status", "Amount (INR)",
        "Attendance Status", "Attended Time"
    ]
    for cf in custom_fields:
        headers.append(cf.field_label)

    writer.writerow(headers)

    for idx, reg in enumerate(registrations, start=1):
        student = reg.student
        profile = student.student_profile
        pay = reg.payment
        att = reg.attendance

        row_data = [
            idx,
            reg.registration_code,
            student.name,
            profile.roll_number if profile else "N/A",
            student.email,
            profile.department if profile else "N/A",
            profile.year if profile else "N/A",
            profile.section if profile else "N/A",
            student.phone or "N/A",
            reg.status,
            pay.status if pay else ("FREE" if event.is_free else "PENDING"),
            event.registration_fee,
            "PRESENT" if att else "ABSENT",
            att.scanned_at.strftime('%Y-%m-%d %H:%M:%S') if att else "N/A"
        ]

        response_map = {r.field_id: r.field_value for r in reg.custom_responses}
        for cf_id in custom_field_ids:
            row_data.append(response_map.get(cf_id, ""))

        writer.writerow(row_data)

    output.seek(0)
    return output
