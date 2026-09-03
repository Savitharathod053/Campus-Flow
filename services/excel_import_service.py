import io
import csv
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email_validator import validate_email, EmailNotValidError

from models import db, User, UserRole, StudentProfile, FacultyProfile

def _normalize_header(header):
    if not header:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(header).strip().lower())

def read_tabular_file(file_storage):
    """
    Reads an uploaded FileStorage object (.xlsx, .xls, or .csv) and returns
    (headers_list, rows_list) as plain text strings.
    """
    filename = file_storage.filename.lower()
    if filename.endswith('.csv'):
        content = file_storage.read().decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(content))
        all_lines = [row for row in reader if any(cell.strip() for cell in row)]
        if not all_lines:
            return [], []
        return all_lines[0], all_lines[1:]
    else:
        file_bytes = io.BytesIO(file_storage.read())
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                all_rows.append([str(cell).strip() if cell is not None else '' for cell in row])
        if not all_rows:
            return [], []
        return all_rows[0], all_rows[1:]


def import_students_excel(file_storage, default_password="Pass@123"):
    """
    Imports student profiles from Excel or CSV.
    Safely creates or updates records within a database transaction.
    """
    headers, rows = read_tabular_file(file_storage)
    if not headers or not rows:
        return {
            'total': 0, 'created': 0, 'updated': 0, 'errors': ['The uploaded file contains no data or empty rows.']
        }

    header_map = {}
    for idx, h in enumerate(headers):
        norm = _normalize_header(h)
        if norm in ('name', 'studentname', 'fullname'):
            header_map['name'] = idx
        elif norm in ('email', 'studentemail', 'collegeemail', 'mail'):
            header_map['email'] = idx
        elif norm in ('phone', 'phonenumber', 'mobile', 'mobilenumber', 'contact'):
            header_map['phone'] = idx
        elif norm in ('rollnumber', 'rollno', 'roll', 'usn', 'registrationno', 'regno'):
            header_map['roll_number'] = idx
        elif norm in ('department', 'dept', 'branch'):
            header_map['department'] = idx
        elif norm in ('year', 'classyear', 'studyyear'):
            header_map['year'] = idx
        elif norm in ('section', 'sec'):
            header_map['section'] = idx
        elif norm in ('password',):
            header_map['password'] = idx

    missing_required = []
    for req in ['name', 'email', 'roll_number', 'department']:
        if req not in header_map:
            missing_required.append(req.replace('_', ' ').title())

    if missing_required:
        return {
            'total': len(rows),
            'created': 0,
            'updated': 0,
            'errors': [f"Missing required column(s): {', '.join(missing_required)}. Expected headers like 'Student Name', 'Email', 'Roll Number', 'Department'."]
        }

    created_count = 0
    updated_count = 0
    errors = []

    try:
        for r_idx, row in enumerate(rows, start=2):
            def get_val(key, default=''):
                pos = header_map.get(key)
                if pos is not None and pos < len(row):
                    return str(row[pos]).strip()
                return default

            name = get_val('name')
            raw_email = get_val('email')
            phone = get_val('phone')
            roll = get_val('roll_number').upper()
            dept = get_val('department').upper()
            year_str = get_val('year', '1')
            section = get_val('section', 'A').upper() or 'A'
            password = get_val('password') or default_password

            if not name or not raw_email or not roll or not dept:
                errors.append(f"Row {r_idx}: Name, Email, Roll Number, and Department cannot be blank.")
                continue

            try:
                valid_info = validate_email(raw_email, check_deliverability=False)
                email = valid_info.normalized.lower()
            except EmailNotValidError as e:
                errors.append(f"Row {r_idx}: Invalid email '{raw_email}' ({str(e)}).")
                continue

            try:
                year = int(float(year_str))
                if year < 1 or year > 6:
                    year = 1
            except ValueError:
                year = 1

            existing_roll = StudentProfile.query.filter_by(roll_number=roll).first()
            existing_user = User.query.filter_by(email=email).first()

            if existing_user:
                if not existing_user.is_student:
                    errors.append(f"Row {r_idx}: User '{email}' exists with conflicting role '{existing_user.role}'. Skipped.")
                    continue

                if existing_roll and existing_roll.user_id != existing_user.id:
                    errors.append(f"Row {r_idx}: Roll number '{roll}' is already assigned to {existing_roll.user.name} ({existing_roll.user.email}).")
                    continue

                existing_user.name = name
                if phone:
                    existing_user.phone = phone
                if not existing_user.student_profile:
                    sp = StudentProfile(
                        user_id=existing_user.id,
                        roll_number=roll,
                        department=dept,
                        year=year,
                        section=section
                    )
                    db.session.add(sp)
                else:
                    existing_user.student_profile.roll_number = roll
                    existing_user.student_profile.department = dept
                    existing_user.student_profile.year = year
                    existing_user.student_profile.section = section

                updated_count += 1
            else:
                if existing_roll:
                    errors.append(f"Row {r_idx}: Roll number '{roll}' is already assigned to {existing_roll.user.name} ({existing_roll.user.email}).")
                    continue

                new_user = User(
                    name=name,
                    email=email,
                    phone=phone or None,
                    role=UserRole.STUDENT,
                    is_active=True
                )
                new_user.set_password(password)
                db.session.add(new_user)
                db.session.flush()

                new_profile = StudentProfile(
                    user_id=new_user.id,
                    roll_number=roll,
                    department=dept,
                    year=year,
                    section=section
                )
                db.session.add(new_profile)
                created_count += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return {
            'total': len(rows),
            'created': 0,
            'updated': 0,
            'errors': [f"Critical database transaction error during import: {str(exc)}"]
        }

    return {
        'total': len(rows),
        'created': created_count,
        'updated': updated_count,
        'errors': errors
    }


def import_faculty_excel(file_storage, default_password="Pass@123"):
    """
    Imports faculty accounts from Excel or CSV.
    Safely creates or updates records within a database transaction.
    """
    headers, rows = read_tabular_file(file_storage)
    if not headers or not rows:
        return {
            'total': 0, 'created': 0, 'updated': 0, 'errors': ['The uploaded file contains no data or empty rows.']
        }

    header_map = {}
    for idx, h in enumerate(headers):
        norm = _normalize_header(h)
        if norm in ('name', 'facultyname', 'fullname', 'professor'):
            header_map['name'] = idx
        elif norm in ('email', 'facultyemail', 'collegeemail', 'mail'):
            header_map['email'] = idx
        elif norm in ('phone', 'phonenumber', 'mobile', 'contact'):
            header_map['phone'] = idx
        elif norm in ('employeeid', 'empid', 'facultyid', 'staffid', 'employeeno'):
            header_map['employee_id'] = idx
        elif norm in ('department', 'dept', 'branch'):
            header_map['department'] = idx
        elif norm in ('designation', 'title', 'role', 'post'):
            header_map['designation'] = idx
        elif norm in ('password',):
            header_map['password'] = idx

    missing_required = []
    for req in ['name', 'email', 'employee_id', 'department']:
        if req not in header_map:
            missing_required.append(req.replace('_', ' ').title())

    if missing_required:
        return {
            'total': len(rows),
            'created': 0,
            'updated': 0,
            'errors': [f"Missing required column(s): {', '.join(missing_required)}. Expected headers like 'Faculty Name', 'Email', 'Employee ID', 'Department'."]
        }

    created_count = 0
    updated_count = 0
    errors = []

    try:
        for r_idx, row in enumerate(rows, start=2):
            def get_val(key, default=''):
                pos = header_map.get(key)
                if pos is not None and pos < len(row):
                    return str(row[pos]).strip()
                return default

            name = get_val('name')
            raw_email = get_val('email')
            phone = get_val('phone')
            emp_id = get_val('employee_id').upper()
            dept = get_val('department').upper()
            desig = get_val('designation') or f"Assistant Professor ({dept})"
            password = get_val('password') or default_password

            if not name or not raw_email or not emp_id or not dept:
                errors.append(f"Row {r_idx}: Name, Email, Employee ID, and Department cannot be blank.")
                continue

            try:
                valid_info = validate_email(raw_email, check_deliverability=False)
                email = valid_info.normalized.lower()
            except EmailNotValidError as e:
                errors.append(f"Row {r_idx}: Invalid email '{raw_email}' ({str(e)}).")
                continue

            existing_emp = FacultyProfile.query.filter_by(employee_id=emp_id).first()
            existing_user = User.query.filter_by(email=email).first()

            if existing_user:
                if not existing_user.is_admin:
                    errors.append(f"Row {r_idx}: User '{email}' exists with conflicting role '{existing_user.role}'. Skipped.")
                    continue

                if existing_emp and existing_emp.user_id != existing_user.id:
                    errors.append(f"Row {r_idx}: Employee ID '{emp_id}' is already assigned to {existing_emp.user.name} ({existing_emp.user.email}).")
                    continue

                existing_user.name = name
                if phone:
                    existing_user.phone = phone
                if not existing_user.faculty_profile:
                    fp = FacultyProfile(
                        user_id=existing_user.id,
                        employee_id=emp_id,
                        department=dept,
                        designation=desig
                    )
                    db.session.add(fp)
                else:
                    existing_user.faculty_profile.employee_id = emp_id
                    existing_user.faculty_profile.department = dept
                    existing_user.faculty_profile.designation = desig

                updated_count += 1
            else:
                if existing_emp:
                    errors.append(f"Row {r_idx}: Employee ID '{emp_id}' is already assigned to {existing_emp.user.name} ({existing_emp.user.email}).")
                    continue

                new_user = User(
                    name=name,
                    email=email,
                    phone=phone or None,
                    role=UserRole.FACULTY_ADMIN,
                    is_active=True
                )
                new_user.set_password(password)
                db.session.add(new_user)
                db.session.flush()

                new_fp = FacultyProfile(
                    user_id=new_user.id,
                    employee_id=emp_id,
                    department=dept,
                    designation=desig
                )
                db.session.add(new_fp)
                created_count += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return {
            'total': len(rows),
            'created': 0,
            'updated': 0,
            'errors': [f"Critical database transaction error during import: {str(exc)}"]
        }

    return {
        'total': len(rows),
        'created': created_count,
        'updated': updated_count,
        'errors': errors
    }


def generate_import_template(target_type="student"):
    """
    Generates a starter Excel (.xlsx) template with sample columns and instructions.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    if target_type == "student":
        ws.title = "Students Import Template"
        headers = ["Student Name", "Email", "Phone", "Roll Number", "Department", "Year", "Section", "Password"]
        sample_rows = [
            ["Aarav Sharma", "aarav.sharma@college.edu", "+91 9876543210", "23CSE01", "CSE", "2", "A", "Pass@123"],
            ["Diya Reddy", "diya.reddy@college.edu", "+91 9876543211", "23ECE02", "ECE", "2", "B", "Pass@123"]
        ]
    else:
        ws.title = "Faculty Import Template"
        headers = ["Faculty Name", "Email", "Phone", "Employee ID", "Department", "Designation", "Password"]
        sample_rows = [
            ["Dr. Arvind Krishnan", "arvind.krishnan@college.edu", "+91 9840112211", "FAC-CSE-201", "CSE", "Professor & Head", "Pass@123"],
            ["Prof. Sunita Rao", "sunita.rao@college.edu", "+91 9840112212", "FAC-ECE-202", "ECE", "Associate Professor", "Pass@123"]
        ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in sample_rows:
        ws.append(r)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
